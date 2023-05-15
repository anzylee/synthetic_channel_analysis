import os
import numpy as np
from matplotlib import cm
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
from scipy.interpolate import interp1d

def flow_num(site_name, flow_condition):
    site_name_all = ['sfe_322', 'sfe_82', 'sfe_4523',
                     'sfe_81', 'sfe_24', 'sfe_25', 'sfe_316',
                     'sfe_2248', 'sfe_221', 'sfe_209']
    num_base = [4, 4, 4, 4,
                4, 1, 4, 4,
                4, 4, 3]
    if flow_condition in ['base']:
        ind = site_name_all.index(site_name)
        num = num_base[ind] # baseflow
    else:
        num = 13 # bankfull
    return num

def remove_zeros_dv(array_d, array_v):
    ind_zero_d = np.where(array_d == 0)
    ind_zero_v = np.where(array_v == 0)

    if ind_zero_d[0].__len__() > ind_zero_v[0].__len__():
        ind_zero = ind_zero_d
    else:
        ind_zero = ind_zero_v
    array_d = np.delete(array_d, ind_zero)
    array_v = np.delete(array_v, ind_zero)
    return array_d, array_v

def histbins_dv(array_d, array_v, method):
    if method in ['Sturges']:
        bin_count_d = int(np.ceil(np.log2(len(array_d))) + 1)
        bin_count_v = int(np.ceil(np.log2(len(array_v))) + 1)
        bin_info = [bin_count_d, bin_count_v]
    elif method in ['fd']:
        q1_d = np.quantile(array_d, 0.25)
        q3_d = np.quantile(array_d, 0.75)
        bin_width_d = (2 * (q3_d - q1_d)) / (len(array_d) ** (1 / 3))
        bin_count_d = int(np.ceil((array_d.max() - array_d.min()) / bin_width_d))

        q1_v = np.quantile(array_v, 0.25)
        q3_v = np.quantile(array_v, 0.75)
        bin_width_v = (2 * (q3_v - q1_v)) / (len(array_v) ** (1 / 3))
        bin_count_v = int(np.ceil((array_v.max() - array_v.min()) / bin_width_v))
        bin_info = [bin_width_d, bin_count_d, bin_width_v, bin_count_v]
    else:
        print()
    return bin_info

def Kullback_div_2d(array_d_TS, array_v_TS, array_d, array_v, method):
    if method in ['Sturges', 'fd']:
        [bin_width_d_TS, bin_count_d_TS, bin_width_v_TS, bin_count_v_TS] = histbins_dv(array_d_TS, array_v_TS, method)
        bin_width_d = bin_width_d_TS
        bin_width_v = bin_width_v_TS
    elif method in ['fixedwidth']:
        bin_width_d = 0.01
        bin_width_v = 0.01
    xedges = np.arange(0, 10, bin_width_d)
    yedges = np.arange(0, 10, bin_width_v)

    H_TS, xedges_TS, yedges_TS = np.histogram2d(array_d_TS, array_v_TS, #bins=[bin_count_d_TS, bin_count_v_TS],
                                                bins=(xedges, yedges), normed=True)
    H, xedges, yedges = np.histogram2d(array_d, array_v, #bins=[bin_count_d, bin_count_v],
                                                bins=(xedges, yedges), normed=True)
    H_TS = H_TS.T
    H = H.T

    H_TS, H = same_size(H_TS, H)

    H_TS[ H_TS == 0 ] = 1e-15
    H[H == 0] = 1e-15

    KLD = H_TS * np.log( H_TS / H ) * xedges[1] * yedges[1]
    # DKL(P||Q) = sum( P(x) * log( P(x) / Q(x) ), P: observed, Q: model
    KLD = np.sum(KLD)

    return KLD, H_TS, xedges_TS, yedges_TS, H, xedges, yedges

def Kullback_div_1d(series_real, series_synthetic, method):
    if method in ['Sturges', 'fd', 'auto']:
        [hist1, edges1] = np.histogram(series_real, bins=method, density=True)
        edges = edges1
        if len(hist1) < 10:
            [hist1, edges1] = np.histogram(series_real, density=True)
            edges = edges1
    elif method in ['fixedwidth']:
        bin_width = 0.5
        edges = np.arange(0, np.max([np.max(series_real), np.max(series_synthetic)]), bin_width)
    elif method in ['default']:
        [hist1, edges1] = np.histogram(series_real, density=True)
        edges = edges1

    [hist1, edges1] = np.histogram(series_real, bins=edges, density=True)
    [hist2, edges2] = np.histogram(series_synthetic, bins=edges, density=True)

    hist1[hist1 == 0] = 1e-15
    hist2[hist2 == 0] = 1e-15

    KLD = hist1 * np.log( hist1 / hist2 ) * np.diff(edges1)  # DKL(P||Q) = sum( P(x) * log( P(x) / Q(x) ), P: observed, Q: model
    KLD = np.sum(KLD)

    return KLD, hist1, edges1, hist2, edges2

def plt_contours(H, xedges, yedges, xlim, ylim, level, curve_type):
    # 2. Contours
    X, Y = np.meshgrid(xedges[1:], yedges[1:])  # to match the size with Z
    colors = ['#FFFFFF', '#F0F0F0', '#D3D3D3', '#B8B8B8']
    fig, ax = plt.subplots()
    CS = ax.contourf(X, Y, H, levels = level, colors=colors, extend = 'both') #, cmap = cm.gray_r
    CS.cmap.set_over('#A0A0A0')

    if curve_type in ['Swaney']:
        periods = ['sp', 'fr', 'ju']
    elif curve_type in ['HTC_ra', 'HTC_co']:
        periods = ['fr', 'ju']
    elif curve_type in ['HTC']:
        periods = ['rafr', 'raju', 'cofr', 'coju']

    if not curve_type in ['false']:
        plt_contours_HSC(curve_type, periods, ax, xlim, ylim)

    plt.xlim([0,xlim])
    plt.ylim([0,ylim])
    plt.xticks(fontsize=18)
    plt.yticks(fontsize=18)
    cbar = fig.colorbar(CS)
    cbar.ax.tick_params(labelsize=18)
    return CS

def plt_contours_HSC(curve_type, periods, ax, xlim, ylim):

    HSC_depth, HSC_velocity = open_HSC(curve_type, periods)
    ind = 0
    colors_line = ['r', 'y', 'b']
    if curve_type in ['Swaney']:
        colors_line = ['#D92014', '#FFe24F', '#1D6FD4']

    elif curve_type in ['HTC_ra']:
        colors_line = ['#4d94ff', '#ff8533']

    elif curve_type in ['HTC_co']:
        colors_line = ['#00b300', '#ffff33']

    elif curve_type in ['HTC']:
        colors_line = ['#4d94ff', '#ff8533', '#00b300', '#ffff33']

    # sp, fr, ju : Red, blue, yellow
    for period in periods:

        var_d_ind = HSC_depth["d"+period] > 0
        var_v_ind = HSC_velocity["v"+period] > 0
        var_d_ind[0] = True
        var_v_ind[0] = True

        f_d = interp1d(HSC_depth["d"+period][var_d_ind], HSC_depth["d"+period+"SI"][var_d_ind])
        f_v = interp1d(HSC_velocity["v"+period][var_v_ind], HSC_velocity["v"+period+"SI"][var_v_ind])

        nx, ny = (50, 50)
        x = np.linspace(0, max(HSC_depth["d"+period][var_d_ind])-0.001, nx)
        y = np.linspace(0, max(HSC_velocity["v"+period][var_v_ind])-0.001, ny)

        X, Y = np.meshgrid(x, y)

        HSI = (f_d(X) * f_v(Y)) ** (1 / 2)

        #ax.contour(X, Y, HSI, levels=[0.5], colors=colors_line[ind])
        CS_HSI = ax.contour(X, Y, HSI, levels=[0.5], colors=(colors_line[ind]))
        CS_HSI = ax.contourf(X, Y, HSI, levels=[0.5, 10], colors=(colors_line[ind], 'w'), alpha = 0.2)
        #CS_HSI.cmap.set_over(colors_line[ind])

        ind += 1

def plt_images(H, xedges, yedges):
    plt.figure()
    plt.imshow(H, interpolation='nearest', origin='lower',
              extent=[xedges[0], xedges[-1], yedges[0], yedges[-1]])
    plt.xlabel('Depth (m)')
    plt.ylabel('Velocity (m)')

def same_size(A, B):
    if A.shape[0] > B.shape[0]:
        rows = A.shape[0] - B.shape[0]
        B = np.append(B, np.zeros((rows, B.shape[1])), axis=0)
    else:
        rows = B.shape[0] - A.shape[0]
        A = np.append(A, np.zeros((rows, A.shape[1])), axis=0)
    if A.shape[1] > B.shape[1]:
        cols = A.shape[1] - B.shape[1]
        B = np.append(B, np.zeros((B.shape[0], cols)), axis=1)
    else:
        cols = B.shape[1] - A.shape[1]
        A = np.append(A, np.zeros((A.shape[0], cols)), axis=1)
    return A, B

def open_HSC(curve_type, periods):
    if curve_type in ['HTC_ra', 'HTC_co']:
        HSC_curve_type = 'HTC'
    else:
        HSC_curve_type = curve_type
    HSC_curve = os.path.join('F:/RiverArchitect-development/.site_packages/templates/fish_'+ HSC_curve_type +'.xlsx')

    wb = openpyxl.Workbook()
    wb = openpyxl.load_workbook(HSC_curve)
    ws = wb['fish']
    if curve_type in ['Swaney']:
        # Depth HSC
        vstart_inds = [9, 9, 9]
        vend_inds = [48, 48, 48]
        # Velocity HSC
        dstart_inds = [51, 51, 51]
        dend_inds = [82, 82, 82]
        cols = ['CD', 'EF', 'GH']
        dcol_names = ['dsp', 'dfr', 'dju', 'dspSI', 'dfrSI', 'djuSI']
        vcol_names = ['vsp', 'vfr', 'vju', 'vspSI', 'vfrSI', 'vjuSI']

    elif curve_type in ['HTC_ra']:
        # Depth HSC
        vstart_inds = [9, 9]
        vend_inds = [110, 110]
        # Velocity HSC
        dstart_inds = [112, 112]
        dend_inds = [211, 211]
        cols = ['EF', 'GH'] # Steelhead Trout
        dcol_names = ['dfr', 'dju', 'dfrSI', 'djuSI']
        vcol_names = ['vfr', 'vju', 'vfrSI', 'vjuSI']

    elif curve_type in ['HTC_co']:
        # Depth HSC
        vstart_inds = [9, 9]
        vend_inds = [110, 110]
        # Velocity HSC
        dstart_inds = [112, 112]
        dend_inds = [211, 211]
        cols = ['MN', 'OP'] # Coho Salmon
        dcol_names = ['dfr', 'dju', 'dfrSI', 'djuSI']
        vcol_names = ['vfr', 'vju', 'vfrSI', 'vjuSI']

    elif curve_type in ['HTC']:
        # Depth HSC
        vstart_inds = [9, 9, 9, 9]
        vend_inds = [110, 110, 110, 110]
        # Velocity HSC
        dstart_inds = [112, 112, 112, 112]
        dend_inds = [211, 211, 211, 211]
        cols = ['EF', 'GH', 'MN', 'OP']  # Coho Salmon
        dcol_names = ['drafr', 'draju', 'dcofr', 'dcoju', 'drafrSI', 'drajuSI', 'dcofrSI', 'dcojuSI']
        vcol_names = ['vrafr', 'vraju', 'vcofr', 'vcoju', 'vrafrSI', 'vrajuSI', 'vcofrSI', 'vcojuSI']

    ind = 0
    tmp_x, tmp_y = [], []

    for period in periods:
        cx = ws[cols[ind][0]+str(dstart_inds[ind]) : cols[ind][0]+str(dend_inds[ind])]
        cy = ws[cols[ind][1] + str(dstart_inds[ind]) : cols[ind][1] + str(dend_inds[ind])]
        ind += 1
        for c1 in cx:
            tmp_x = np.append(tmp_x, c1[0].value)
        for c2 in cy:
            tmp_y = np.append(tmp_y, c2[0].value)

    tmp = np.array([tmp_x, tmp_y])
    tmp = np.where(tmp==None, 0, tmp)
    tmp = tmp.T

    #print(len(tmp))
    tmp = tmp.reshape(int(len(tmp) / len(cols)), int(len(dcol_names)), order='F')

    HSC_depth =pd.DataFrame(tmp, columns = dcol_names)

    ind = 0
    tmp_x, tmp_y = [], []

    for period in periods:
        cx = ws[cols[ind][0]+str(vstart_inds[ind]) : cols[ind][0]+str(vend_inds[ind])]
        cy = ws[cols[ind][1] + str(vstart_inds[ind]) : cols[ind][1] + str(vend_inds[ind])]
        ind += 1
        for c1 in cx:
            tmp_x = np.append(tmp_x, c1[0].value)
        for c2 in cy:
            tmp_y = np.append(tmp_y, c2[0].value)

    tmp = np.array([tmp_x, tmp_y])
    tmp = np.where(tmp==None, 0, tmp)
    tmp = tmp.T

    tmp = tmp.reshape(int(len(tmp) / len(cols)), int(len(vcol_names)), order='F')

    HSC_velocity =pd.DataFrame(tmp, columns = vcol_names)

    # ft to m
    for column in HSC_depth:
        if not column[-2:] == 'SI':
            HSC_depth[column] = HSC_depth[column]*0.3048
    for column in HSC_velocity:
        if not column[-2:] == 'SI':
            HSC_velocity[column] = HSC_velocity[column]*0.3048

    return HSC_depth, HSC_velocity